from mpl_toolkits.mplot3d import Axes3D
import openpyxl
import numpy as np
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook('finalML.xlsx')
ws = wb.get_sheet_by_name("Sheet1")

xlist = []
ylist = []
zlist = []
flist = []
for x in range(1, 1271):
    d = ws.cell(row=x, column=2).value
    e = ws.cell(row=x, column=4).value
    f = ws.cell(row=x, column=5).value
    g = ws.cell(row=x, column=6).value

    xlist.append(float(d))
    ylist.append(float(e))
    zlist.append(f)

fig = plt.figure()

ax = fig.add_subplot(111, projection='3d')

x = np.array([xlist])
y = np.array([ylist])
z = np.array([zlist])

flist = [1, 0, 0, 0, 1, 1, 1, 0, 0, 1,
         0, 1, 1, 0, 1, 1, 0, 0, 0, 1,
         0, 1, 1, 1, 1, 1, 0, 0, 0, 0,
         0, 0, 0, 1, 1, 1, 1, 1, 1, 1,
         0, 0, 1, 1, 1, 1, 1, 0, 0, 1,
         1, 1, 1, 1, 0, 0, 1, 1, 1, 1,
         1, 0, 0, 1, 1, 0, 1, 1, 1, 0,
         1, 1, 0, 1, 1, 0, 1, 1, 0, 1,
         0, 0, 1, 0, 1, 1, 0, 1, 0, 0,
         1, 1, 0, 0, 0, 0, 0, 1, 0, 1,
         1, 1, 1, 0, 1, 0, 0, 1, 0, 0,
         1, 0, 0, 1, 0, 1, 1, 1, 1, 1,
         1, 0, 0, 1, 1, 0, 1, 1, 1, 1,
         0, 1, 1, 1, 0, 1, 1, 1, 1, 0,
         1, 1, 1, 1, 1, 1, 1, 1, 0, 1,
         0, 0, 0, 1, 1, 1, 1, 1, 1, 0,
         1, 1, 0, 0, 1, 1, 1, 0, 1, 0,
         1, 1, 0, 1, 1, 0, 1, 1, 1, 0,
         1, 1, 1, 1, 0, 1, 0, 1, 0, 1,
         1, 1, 1, 0, 1, 1, 1, 0, 1, 1,
         1, 1, 0, 1, 0, 1, 1, 1, 0, 1,
         1, 1, 1, 1, 1, 1, 1, 1, 1, 0,
         1, 1, 0, 1, 1, 1, 1, 0, 1, 1,
         0, 0, 0, 1, 0, 0, 1, 0, 1, 1,
         1, 1, 0, 1, 1, 1, 1, 0, 0, 0,
         1, 1, 0, 0, 0, 0, 1, 1, 1, 1,
         0, 1, 0, 1, 0, 1, 1, 0, 1, 1,
         1, 1, 1, 1, 0, 1, 1, 0, 0, 1,
         1, 1, 1, 0, 1, 1, 0, 1, 1, 1,
         1, 1, 1, 1, 1, 1, 0, 0, 1, 1,
         1, 1, 1, 1, 1, 1, 1, 0, 0, 0,
         1, 0, 1, 1, 0, 1, 1, 1, 0, 0,
         0, 1, 0, 1, 0, 1, 1, 0, 1, 0,
         1, 0, 0, 1, 0, 1, 0, 1, 1, 1,
         1, 1, 0, 0, 1, 1, 0, 0, 1, 1,
         1, 0, 0, 0, 1, 1, 0, 1, 1, 1,
         1, 1, 1, 0, 1, 1, 1, 1, 1, 1,
         1, 0, 1, 1, 0, 0, 0, 1, 1, 1,
         1, 1, 1, 1, 1, 0, 0, 1, 1, 0,
         1, 1, 1, 1, 1, 1, 1, 0, 0, 1,
         1, 1, 1, 1, 0, 0, 0, 1, 0, 1,
         0, 0, 0, 1, 1, 0, 1, 1, 1, 0,
         1, 1, 0, 1, 1, 0, 1, 1, 1, 0,
         1, 1, 0, 1, 0, 1, 0, 0, 1, 1,
         1, 1, 1, 0, 0, 1, 0, 0, 1, 1,
         0, 1, 1, 0, 1, 0, 1, 1, 1, 1,
         0, 1, 0, 1, 0, 1, 0, 0, 0, 0,
         0, 1, 1, 1, 1, 1, 1, 1, 1, 1,
         1, 1, 1, 1, 0, 0, 0, 0, 0, 1,
         0, 1, 1, 0, 1, 0, 1, 1, 0, 1,
         1, 0, 0, 0, 1, 0, 0, 0, 0, 1,
         0, 0, 1, 0, 1, 0, 1, 1, 1, 0,
         1, 0, 0, 1, 0, 0, 1, 0, 0, 1,
         0, 1, 1, 0, 1, 1, 0, 1, 1, 1,
         1, 1, 1, 0, 1, 0, 0, 0, 0, 1,
         0, 1, 1, 0, 1, 1, 0, 0, 1, 1,
         1, 0, 1, 1, 0, 1, 1, 0, 1, 0,
         1, 1, 1, 1, 1, 0, 0, 0, 1, 1,
         1, 1, 1, 1, 0, 1, 0, 0, 0, 0,
         1, 1, 0, 0, 1, 0, 0, 1, 1, 1,
         0, 0, 0, 0, 1, 0, 0, 0, 1, 1,
         1, 1, 0, 1, 0, 0, 1, 1, 0, 1,
         0, 0, 1, 1, 1, 0, 0, 1, 0, 1,
         1, 0, 0, 1, 0, 0, 1, 0, 1, 1,
         1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
         1, 0, 0, 0, 0, 1, 0, 1, 0, 1,
         1, 1, 1, 0, 1, 0, 1, 0, 1, 1,
         0, 1, 0, 0, 1, 1, 0, 0, 0, 0,
         1, 1, 0, 1, 0, 1, 0, 0, 1, 1,
         0, 1, 1, 0, 1, 0, 1, 0, 0, 0,
         1, 0, 0, 1, 0, 1, 1, 1, 1, 1,
         1, 1, 1, 1, 1, 0, 1, 1, 0, 0,
         1, 1, 1, 0, 0, 0, 0, 1, 0, 1,
         0, 0, 1, 0, 1, 0, 0, 0, 0, 0,
         1, 0, 1, 0, 1, 1, 1, 0, 0, 1,
         1, 1, 1, 1, 0, 1, 1, 1, 0, 1,
         0, 0, 1, 1, 0, 0, 0, 0, 0, 1,
         1, 1, 0, 1, 0, 1, 1, 1, 1, 1,
         0, 1, 0, 0, 1, 1, 1, 1, 1, 0,
         1, 0, 1, 1, 0, 1, 1, 1, 1, 1,
         1, 0, 0, 1, 1, 0, 0, 0, 1, 0,
         1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
         1, 1, 1, 1, 1, 1, 0, 1, 0, 0,
         1, 1, 0, 1, 0, 1, 1, 0, 1, 0,
         0, 1, 1, 1, 1, 0, 0, 0, 1, 0,
         0, 0, 1, 0, 1, 1, 0, 1, 0, 0,
         1, 0, 1, 0, 1, 0, 1, 1, 0, 0,
         0, 1, 1, 1, 0, 1, 1, 1, 0, 0,
         1, 0, 1, 0, 1, 0, 1, 0, 0, 1,
         1, 1, 0, 0, 1, 1, 0, 1, 0, 1,
         1, 0, 1, 0, 0, 1, 0, 0, 1, 0,
         1, 1, 0, 0, 1, 1, 1, 1, 0, 1,
         0, 1, 1, 0, 0, 1, 0, 0, 1, 1,
         1, 1, 0, 1, 1, 0, 1, 0, 0, 1,
         1, 0, 1, 0, 0, 1, 1, 1, 0, 1,
         1, 0, 1, 1, 0, 0, 1, 1, 0, 0,
         0, 1, 1, 1, 1, 1, 1, 1, 1, 0,
         1, 1, 0, 1, 0, 0, 1, 0, 0, 1,
         0, 0, 1, 1, 1, 0, 1, 1, 0, 0,
         1, 1, 1, 0, 1, 0, 0, 1, 0, 0,
         1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
         1, 0, 1, 0, 0, 0, 0, 0, 1, 1,
         1, 1, 0, 1, 1, 1, 1, 1, 1, 1,
         1, 1, 0, 1, 0, 1, 1, 1, 1, 0,
         1, 0, 1, 1, 1, 1, 1, 0, 0, 0,
         1, 0, 1, 0, 0, 0, 1, 0, 0, 1,
         1, 1, 0, 0, 0, 1, 0, 1, 1, 0,
         0, 0, 0, 1, 0, 0, 1, 0, 1, 1,
         0, 1, 0, 1, 0, 1, 0, 1, 1, 0,
         1, 0, 1, 1, 0, 1, 1, 0, 0, 1,
         0, 1, 1, 1, 0, 1, 1, 0, 1, 1,
         1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
         1, 0, 1, 1, 1, 0, 0, 1, 0, 1,
         1, 0, 1, 0, 1, 0, 1, 1, 0, 1,
         1, 0, 1, 0, 1, 0, 0, 1, 0, 0,
         0, 1, 1, 0, 1, 1, 0, 1, 0, 0,
         0, 0, 1, 0, 1, 0, 1, 1, 1, 1,
         0, 1, 0, 1, 1, 1, 1, 1, 1, 0,
         1, 1, 1, 1, 1, 1, 1, 1, 0, 1,
         0, 1, 1, 1, 1, 1, 1, 0, 0, 1,
         1, 1, 0, 0, 1, 1, 0, 1, 1, 0,
         1, 1, 1, 0, 1, 1, 0, 1, 0, 1,
         1, 0, 1, 0, 1, 0, 1, 1, 0, 0,
         1, 0, 1, 1, 1, 1, 0, 0, 0, 1,
         0, 0, 1, 1, 1, 1, 1, 1, 1, 1,
         1, 1, 1, 0, 0, 0, 1, 1, 1, 1,
         0, 0, 1, 0, 1, 0, 1, 1, 1, 1]

ax.scatter(x, y, z, c=flist, marker='o')

ax.set_xlabel('X Label--longitude')
ax.set_ylabel('Y Label--cloud coverage')
ax.set_zlabel('Z Label--elevation')
plt.show()

